# -*- coding: utf-8 -*-
"""
aicp_goodput_max_concurrency_sweep.py
======================================

对 **已运行** 的 SGLang 服务端，以随机 token 输入（``ignore_eos=True``）扫描
不同 seq-len 组合下，对每个 goodput SLO key **分别**求出满足 100% 的最大并发数。

追踪的 key 集合（由配置的 ``goodput`` SLO 决定）：
  * ``combined``   —— 所有 SLO key 同时满足（goodput_percentage == 100%）
  * ``ttft``       —— 仅 TTFT SLO 满足（ttft_goodput_pct == 100%）
  * ``tpot``       —— 仅 TPOT SLO 满足
  * ``e2el``       —— 仅 E2EL SLO 满足
  * ``throughput`` —— 仅 throughput SLO 满足
  （仅配置了的 key 才会追踪）

搜索策略（步进式，非二分）：
  并发序列：1 → 2 → 4 → 8 → 12 → 16 → 20 → …（前段×2，之后 +linear_step）
  • 对每个并发点，检查所有仍 "active" 的 key
  • 某 key 第一次不满足 100%，则记录其最大并发 = 上一个通过的并发，并移出 active
  • 所有 key 均已确定后停止搜索（无需跑到 max_batch_size）

输出：
  1. partial JSONL（边跑边追加，崩溃不丢数据）
  2. 每个 seqlen 一张 **详情 Excel**：行 = 每个测试并发点，列 = 全量指标
  3. 最终 **汇总 JSONL**：每行一条 seqlen，含各 key 的最大并发
  4. 最终 **汇总 Excel**：第一列 seqlen，其余列 = 各 key 最大并发（+ 关键指标）

访问方法与 ``aicp_random_tokens_benchmark.py`` 完全一致：
  * /v1/chat/completions 流式接口 + ignore_eos=True
  * asyncio.Semaphore 精确控制并发
  * 同款 TTFT/TPOT/ITL/E2EL/吞吐 + per-key goodput 指标

配置文件（JSON）字段：

  base_url               服务端 URL，例如 http://127.0.0.1:30000
  min_input_len          扫描起始输入长度（token）
  max_input_len          扫描结束输入长度（token）
  seqlen_step            输入长度步长
  output_len             固定输出长度（max_tokens，ignore_eos=True）
  combo_per_batch_size   每个并发点 num_prompts = batch_size × combo（默认 3）
  min_batch_size         并发搜索起点（默认 1）
  max_batch_size         并发搜索上限（默认 64）
  linear_step            步进阶段的步长（默认 4，即 8→12→16→…）
  goodput                SLO 列表，格式 "KEY:VALUE"，支持 ttft/tpot/e2el/throughput
                         例如 ["ttft:5000", "throughput:10"]
  model                  模型名（不填则 GET /v1/models 自动探测）
  api_key                Bearer token（默认 EMPTY）
  prompt_mode            "approx" 或 "exact"（默认 approx）
  tokenizer_path         exact 模式必填
  seed                   随机种子（默认 1）
  output_path            汇总 JSONL 路径（默认 goodput_max_concurrency_results.jsonl）
  detail_excel_dir       每个 seqlen 详情 Excel 的目录（默认 null = 当前目录）
  disable_excel          不生成任何 Excel（默认 false）
  excel_path             汇总 Excel 路径（不填则自动命名）
  warmup_input_len       预热输入长度（默认 32）
  warmup_output_len      预热输出长度（默认 32）
  disable_warmup         跳过预热（默认 false）

用法：

  python benchmark/speculative_aicp/aicp_goodput_max_concurrency_sweep.py \\
      --config benchmark/speculative_aicp/goodput_sweep_config.example.json
"""

from __future__ import annotations

import argparse
import asyncio
import json
import logging
import os
import random
import re
import sys
import time
import traceback
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, AsyncGenerator, Dict, List, Optional, Set, Tuple

import aiohttp
import numpy as np
import requests

try:
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.styles import Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.utils.dataframe import dataframe_to_rows
    _EXCEL_AVAILABLE = True
except ImportError:
    _EXCEL_AVAILABLE = False

try:
    from tqdm.asyncio import tqdm as async_tqdm
    _TQDM_AVAILABLE = True
except ImportError:
    _TQDM_AVAILABLE = False

try:
    from transformers import AutoTokenizer
    _TRANSFORMERS_AVAILABLE = True
except ImportError:
    _TRANSFORMERS_AVAILABLE = False
    AutoTokenizer = None  # type: ignore[assignment]


###############################################################################
# Constants
###############################################################################

MILLISECONDS_TO_SECONDS_CONVERSION = 1000
AIOHTTP_TIMEOUT = aiohttp.ClientTimeout(total=6 * 60 * 60)
_SELECTED_PERCENTILES = [90.0, 95.0, 99.0]


###############################################################################
# Data structures (aligned with aicp_random_tokens_benchmark.py)
###############################################################################

@dataclass
class BenchmarkMetrics:
    completed: int
    mean_input: int
    mean_output: int
    total_input: int
    total_output: int
    request_throughput: float
    request_goodput: float
    goodput_percentage: float
    output_throughput: float
    mean_output_throughput: float
    total_token_throughput: float
    mean_goodput_ttft: float
    mean_goodput_tpot: float
    mean_goodput_e2el: float
    mean_goodput_throughput: float
    mean_ttft_ms: float
    median_ttft_ms: float
    std_ttft_ms: float
    percentiles_ttft_ms: List[Tuple[float, float]]
    mean_tpot_ms: float
    median_tpot_ms: float
    std_tpot_ms: float
    percentiles_tpot_ms: List[Tuple[float, float]]
    mean_itl_ms: float
    median_itl_ms: float
    std_itl_ms: float
    percentiles_itl_ms: List[Tuple[float, float]]
    mean_e2el_ms: float
    median_e2el_ms: float
    std_e2el_ms: float
    percentiles_e2el_ms: List[Tuple[float, float]]
    mean_prefilling_throughput: float
    median_prefilling_throughput: float
    std_prefilling_throughput: float
    percentiles_prefilling_throughput: List[Tuple[float, float]]
    mean_decoding_throughput: float
    median_decoding_throughput: float
    std_decoding_throughput: float
    percentiles_decoding_throughput: List[Tuple[float, float]]


@dataclass
class RequestFuncInput:
    prompt: str
    api_url: str
    prompt_len: int
    output_len: int
    model: str
    best_of: int = 1
    logprobs: Optional[int] = None
    multi_modal_content: Optional[dict] = None
    ignore_eos: bool = False
    temperature: float = 0.0
    stream: bool = True


@dataclass
class RequestFuncOutput:
    generated_text: str = ""
    success: bool = False
    latency: float = 0.0
    ttft: float = 0.0
    itl: List[float] = field(default_factory=list)
    prompt_len: int = 0
    completion_len: int = 0
    error: str = ""


###############################################################################
# Config
###############################################################################

@dataclass
class SweepConfig:
    base_url: str

    # Seqlen sweep
    min_input_len: int = 500
    max_input_len: int = 5000
    seqlen_step: int = 500
    output_len: int = 256

    # Concurrency search
    min_batch_size: int = 1
    max_batch_size: int = 64
    linear_step: int = 4            # step size after the initial doubling phase

    # Benchmark
    combo_per_batch_size: int = 3
    goodput: List[str] = field(default_factory=lambda: ["ttft:5000", "throughput:10"])

    # Server / model
    model: Optional[str] = None
    api_key: str = "EMPTY"
    prompt_mode: str = "approx"
    tokenizer_path: Optional[str] = None
    seed: int = 1

    # Warmup
    warmup_input_len: int = 32
    warmup_output_len: int = 32
    disable_warmup: bool = False

    # Output
    output_path: str = "goodput_max_concurrency_results.jsonl"
    detail_excel_dir: Optional[str] = None   # per-seqlen detail Excel dir; None = cwd
    disable_excel: bool = False
    excel_path: Optional[str] = None         # final summary Excel path

    @classmethod
    def from_file(cls, path: str) -> "SweepConfig":
        with open(path, "r", encoding="utf-8") as f:
            raw = json.load(f)
        raw = {k: v for k, v in raw.items() if not k.startswith("_")}
        unknown = set(raw) - set(cls.__dataclass_fields__)
        if unknown:
            raise ValueError(f"Unknown config keys: {sorted(unknown)}")
        obj = cls(**raw)
        obj.validate()
        return obj

    def validate(self) -> None:
        if self.min_input_len <= 0 or self.seqlen_step <= 0:
            raise ValueError("min_input_len and seqlen_step must be > 0")
        if self.max_input_len < self.min_input_len:
            raise ValueError("max_input_len must be >= min_input_len")
        if self.output_len <= 1:
            raise ValueError("output_len must be > 1 (needed for TPOT/ITL)")
        if self.min_batch_size < 1:
            raise ValueError("min_batch_size must be >= 1")
        if self.max_batch_size < self.min_batch_size:
            raise ValueError("max_batch_size must be >= min_batch_size")
        if self.linear_step < 1:
            raise ValueError("linear_step must be >= 1")
        if self.combo_per_batch_size < 1:
            raise ValueError("combo_per_batch_size must be >= 1")
        if self.prompt_mode not in ("approx", "exact"):
            raise ValueError("prompt_mode must be 'approx' or 'exact'")
        if self.prompt_mode == "exact" and not self.tokenizer_path:
            raise ValueError("tokenizer_path is required when prompt_mode='exact'")

    def seqlen_list(self) -> List[int]:
        return list(range(self.min_input_len, self.max_input_len + 1, self.seqlen_step))


###############################################################################
# Benchmark core helpers (aligned with aicp_random_tokens_benchmark.py)
###############################################################################

def remove_prefix(s: str, prefix: str) -> str:
    if s.startswith(prefix):
        return s[len(prefix):]
    return s


def parse_goodput(slo_pairs: List[str]) -> Dict[str, float]:
    out: Dict[str, float] = {}
    for pair in slo_pairs:
        name, val = pair.split(":", 1)
        out[name.strip()] = float(val.strip())
    return out


def check_goodput_args(slo_args: Optional[List[str]]) -> Dict[str, float]:
    if not slo_args:
        return {}
    VALID_NAMES = ["ttft", "tpot", "e2el", "throughput", "ttft2", "throughput2"]
    cfg = parse_goodput(slo_args)
    for name, val in cfg.items():
        if name not in VALID_NAMES:
            raise ValueError(f"Invalid goodput metric: {name!r}")
        if val < 0:
            raise ValueError(f"Invalid goodput value: {name}:{val}")
    return cfg


def calculate_metrics(
    outputs: Tuple[RequestFuncOutput, ...],
    dur_s: float,
    selected_percentiles: List[float],
    goodput_config_dict: Dict[str, float],
) -> Tuple[BenchmarkMetrics, List[int], Dict[str, float]]:
    actual_output_lens: List[int] = []
    total_input = 0
    completed = 0
    good_completed = 0
    itls: List[float] = []
    tpots: List[float] = []
    all_tpots: List[float] = []
    ttfts: List[float] = []
    e2els: List[float] = []
    out_throughputs: List[float] = []
    prefill_throughputs: List[float] = []
    decode_throughputs: List[float] = []

    for i in range(len(outputs)):
        if outputs[i].success:
            output_len = outputs[i].completion_len
            actual_output_lens.append(output_len)
            total_input += outputs[i].prompt_len
            tpot, throughput = 0.0, 0.0
            if output_len > 1:
                tpot = (outputs[i].latency - outputs[i].ttft) / (output_len - 1)
                tpots.append(tpot)
                throughput = output_len / outputs[i].latency
                out_throughputs.append(throughput)
                prefill_throughputs.append(outputs[i].prompt_len / outputs[i].ttft)
                decode_throughputs.append(
                    (output_len - 1) / (outputs[i].latency - outputs[i].ttft)
                )
            all_tpots.append(tpot)
            itls += outputs[i].itl
            ttfts.append(outputs[i].ttft)
            e2els.append(outputs[i].latency)
            completed += 1
        else:
            actual_output_lens.append(0)

    mean_goodput_ttft = float("nan")
    mean_goodput_tpot = float("nan")
    mean_goodput_e2el = float("nan")
    mean_goodput_throughput = float("nan")
    per_key_counts: Dict[str, int] = {}
    n_zip = 0

    if goodput_config_dict:
        slo_values: Dict[str, float] = {
            "ttft": (
                goodput_config_dict["ttft"] / MILLISECONDS_TO_SECONDS_CONVERSION
                if "ttft" in goodput_config_dict else float("inf")
            ),
            "tpot": (
                goodput_config_dict["tpot"] / MILLISECONDS_TO_SECONDS_CONVERSION
                if "tpot" in goodput_config_dict else float("inf")
            ),
            "e2el": (
                goodput_config_dict["e2el"] / MILLISECONDS_TO_SECONDS_CONVERSION
                if "e2el" in goodput_config_dict else float("inf")
            ),
            "throughput": (
                goodput_config_dict["throughput"]
                if "throughput" in goodput_config_dict else float("-inf")
            ),
        }

        good_ttfts = good_tpots = good_e2els = good_throughputs = np.array([])
        per_key_counts = {k: 0 for k in goodput_config_dict}

        for _ttft, _tpot, _e2el, _throughput in zip(ttfts, all_tpots, e2els, decode_throughputs):
            n_zip += 1
            if (
                _ttft <= slo_values["ttft"]
                and _tpot <= slo_values["tpot"]
                and _e2el <= slo_values["e2el"]
                and _throughput >= slo_values["throughput"]
            ):
                good_ttfts = np.append(good_ttfts, _ttft)
                good_tpots = np.append(good_tpots, _tpot)
                good_e2els = np.append(good_e2els, _e2el)
                good_throughputs = np.append(good_throughputs, _throughput)
            if "ttft" in goodput_config_dict and _ttft <= slo_values["ttft"]:
                per_key_counts["ttft"] += 1
            if "tpot" in goodput_config_dict and _tpot <= slo_values["tpot"]:
                per_key_counts["tpot"] += 1
            if "e2el" in goodput_config_dict and _e2el <= slo_values["e2el"]:
                per_key_counts["e2el"] += 1
            if "throughput" in goodput_config_dict and _throughput >= slo_values["throughput"]:
                per_key_counts["throughput"] += 1

        good_completed = len(good_ttfts)
        if good_completed > 0:
            mean_goodput_ttft = float(np.mean(good_ttfts))
            mean_goodput_tpot = float(np.mean(good_tpots))
            mean_goodput_e2el = float(np.mean(good_e2els))
            mean_goodput_throughput = float(np.mean(good_throughputs))

    if completed == 0:
        logging.warning("All requests failed. Check server URL, model name, and GPU health.")
        completed = float("-inf")  # type: ignore[assignment]

    def _pct(seq: List[float], p: float) -> float:
        return float(np.percentile(seq or [0], p))

    metrics = BenchmarkMetrics(
        completed=completed,
        mean_input=int(total_input / completed) if completed > 0 else 0,
        mean_output=int(np.mean(actual_output_lens) if actual_output_lens else 0),
        total_input=total_input,
        total_output=sum(actual_output_lens),
        request_throughput=completed / dur_s if completed > 0 else 0.0,
        request_goodput=good_completed / dur_s,
        goodput_percentage=(good_completed / completed) * 100 if completed > 0 else 0.0,
        output_throughput=sum(actual_output_lens) / dur_s,
        mean_output_throughput=float(np.mean(out_throughputs or [0])),
        total_token_throughput=(total_input + sum(actual_output_lens)) / dur_s,
        mean_goodput_ttft=float(mean_goodput_ttft),
        mean_goodput_tpot=float(mean_goodput_tpot),
        mean_goodput_e2el=float(mean_goodput_e2el),
        mean_goodput_throughput=float(mean_goodput_throughput),
        mean_ttft_ms=float(np.mean(ttfts or [0])) * 1000,
        std_ttft_ms=float(np.std(ttfts or [0])) * 1000,
        median_ttft_ms=float(np.median(ttfts or [0])) * 1000,
        percentiles_ttft_ms=[(p, _pct(ttfts, p) * 1000) for p in selected_percentiles],
        mean_tpot_ms=float(np.mean(tpots or [0])) * 1000,
        std_tpot_ms=float(np.std(tpots or [0])) * 1000,
        median_tpot_ms=float(np.median(tpots or [0])) * 1000,
        percentiles_tpot_ms=[(p, _pct(tpots, p) * 1000) for p in selected_percentiles],
        mean_itl_ms=float(np.mean(itls or [0])) * 1000,
        std_itl_ms=float(np.std(itls or [0])) * 1000,
        median_itl_ms=float(np.median(itls or [0])) * 1000,
        percentiles_itl_ms=[(p, _pct(itls, p) * 1000) for p in selected_percentiles],
        mean_e2el_ms=float(np.mean(e2els or [0])) * 1000,
        std_e2el_ms=float(np.std(e2els or [0])) * 1000,
        median_e2el_ms=float(np.median(e2els or [0])) * 1000,
        percentiles_e2el_ms=[(p, _pct(e2els, p) * 1000) for p in selected_percentiles],
        mean_prefilling_throughput=float(np.mean(prefill_throughputs or [0])),
        std_prefilling_throughput=float(np.std(prefill_throughputs or [0])),
        median_prefilling_throughput=float(np.median(prefill_throughputs or [0])),
        percentiles_prefilling_throughput=[
            (p, float(_pct(prefill_throughputs, p))) for p in selected_percentiles
        ],
        mean_decoding_throughput=float(np.mean(decode_throughputs or [0])),
        std_decoding_throughput=float(np.std(decode_throughputs or [0])),
        median_decoding_throughput=float(np.median(decode_throughputs or [0])),
        percentiles_decoding_throughput=[
            (p, float(_pct(decode_throughputs, p))) for p in selected_percentiles
        ],
    )

    per_key_goodput_pct: Dict[str, float] = {}
    if goodput_config_dict and n_zip > 0:
        per_key_goodput_pct = {k: cnt / n_zip * 100 for k, cnt in per_key_counts.items()}
    elif goodput_config_dict:
        per_key_goodput_pct = {k: 0.0 for k in goodput_config_dict}

    return metrics, actual_output_lens, per_key_goodput_pct


async def async_request_openai(
    request_func_input: RequestFuncInput,
    api_key: str,
    enable_thinking: bool = False,
    pbar: Optional[Any] = None,
) -> RequestFuncOutput:
    api_url = request_func_input.api_url
    connector = aiohttp.TCPConnector(ssl=False) if "https" in api_url else None
    async with aiohttp.ClientSession(timeout=AIOHTTP_TIMEOUT, connector=connector) as session:
        payload: Dict[str, Any] = {
            "model": request_func_input.model,
            "temperature": request_func_input.temperature,
            "best_of": request_func_input.best_of,
            "max_tokens": request_func_input.output_len,
            "stream": request_func_input.stream,
            "ignore_eos": request_func_input.ignore_eos,
            "stream_options": {"include_usage": True},
            "chat_template_kwargs": {"enable_thinking": enable_thinking},
            "messages": [{"role": "user", "content": request_func_input.prompt}],
        }
        headers = {
            "Authorization": f"Bearer {api_key}",
            "Content-Type": "application/json",
        }
        output = RequestFuncOutput()
        output.prompt_len = request_func_input.prompt_len
        st = time.perf_counter()
        most_recent_timestamp = st
        try:
            async with session.post(url=api_url, json=payload, headers=headers) as response:
                if response.status == 200:
                    async for chunk_bytes in response.content:
                        if (
                            not chunk_bytes
                            or chunk_bytes in {b"\r", b"\n"}
                            or b": ping - " in chunk_bytes
                        ):
                            continue
                        if b"local_rate_limited" in chunk_bytes:
                            break
                        chunk = remove_prefix(
                            chunk_bytes.decode("utf-8"), "data:"
                        ).strip()
                        if chunk == "[DONE]":
                            output.latency = time.perf_counter() - st
                            output.success = True
                            if output.completion_len == 0 and output.itl:
                                output.completion_len = len(output.itl) + 1
                        else:
                            timestamp = time.perf_counter()
                            data = json.loads(chunk)
                            if data.get("status", {}).get("code") == 500:
                                output.error = f"{response.reason} {chunk}".strip()
                                output.success = False
                                break
                            choices = data.get("choices", [])
                            if not choices and data.get("usage"):
                                output.prompt_len = data["usage"]["prompt_tokens"]
                                output.completion_len = data["usage"]["completion_tokens"]
                                continue
                            if not choices:
                                continue
                            delta = choices[0].get("delta", {})
                            if "content" in delta:
                                if not delta.get("content"):
                                    continue
                                if output.ttft == 0.0:
                                    output.ttft = time.perf_counter() - st
                                else:
                                    output.itl.append(timestamp - most_recent_timestamp)
                                output.generated_text += delta["content"]
                                most_recent_timestamp = timestamp
                else:
                    error_message = "".join(
                        [c.decode("utf-8").strip() async for c in response.content]
                    )
                    output.error = f"{response.reason} {error_message}".strip()
                    output.success = False
        except Exception:
            output.success = False
            output.error = "".join(traceback.format_exception(*sys.exc_info()))
            logging.error("Request failed: %s", output.error)

    if pbar is not None:
        pbar.update(1)
    return output


async def limited_request_func(
    semaphore: Optional[asyncio.Semaphore],
    request_func_input: RequestFuncInput,
    api_key: str,
    enable_thinking: bool,
    pbar: Optional[Any],
) -> RequestFuncOutput:
    if semaphore is None:
        return await async_request_openai(
            request_func_input, api_key=api_key,
            enable_thinking=enable_thinking, pbar=pbar,
        )
    async with semaphore:
        return await async_request_openai(
            request_func_input, api_key=api_key,
            enable_thinking=enable_thinking, pbar=pbar,
        )


async def get_request(
    input_requests: List[Tuple[str, int, int, Optional[dict]]],
    request_rate: float,
) -> AsyncGenerator[Tuple[str, int, int, Optional[dict]], None]:
    for request in input_requests:
        yield request
        if request_rate != float("inf"):
            await asyncio.sleep(1.0 / request_rate)


def build_approx_prompt_tokens(seqlen: int, rng: random.Random) -> str:
    return " ".join(str(rng.randint(0, 99)) for _ in range(seqlen))


def build_exact_prompt_tokens(
    tokenizer: Any,
    num_tokens: int,
    rng: random.Random,
) -> Tuple[str, int]:
    if num_tokens <= 0:
        raise ValueError("num_tokens must be > 0")
    ids: List[int] = []
    special = set(getattr(tokenizer, "all_special_ids", []) or [])
    for _ in range(10000):
        while len(ids) < num_tokens:
            tid = rng.randint(0, tokenizer.vocab_size - 1)
            if tid not in special:
                ids.append(tid)
        ids = ids[:num_tokens]
        text = tokenizer.decode(ids, skip_special_tokens=True)
        enc = tokenizer.encode(text, add_special_tokens=False)
        if len(enc) == num_tokens:
            return text, num_tokens
        ids = enc[:num_tokens] if len(enc) > num_tokens else enc
    raise RuntimeError(f"无法构造恰好 {num_tokens} 个 token 的 prompt。")


def make_input_requests(
    *,
    prompt_mode: str,
    tokenizer: Optional[Any],
    num_prompts: int,
    input_len: int,
    output_len: int,
    seed: int,
) -> List[Tuple[str, int, int, None]]:
    rng = random.Random(seed)
    out: List[Tuple[str, int, int, None]] = []
    for _ in range(num_prompts):
        if prompt_mode == "exact":
            assert tokenizer is not None
            prompt, plen = build_exact_prompt_tokens(tokenizer, input_len, rng)
            out.append((prompt, plen, output_len, None))
        else:
            prompt = build_approx_prompt_tokens(input_len, rng)
            out.append((prompt, input_len, output_len, None))
    return out


def fetch_model_name(base_url: str, api_key: str) -> Optional[str]:
    headers = {"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"}
    try:
        r = requests.get(base_url.rstrip("/") + "/v1/models", headers=headers, timeout=30)
        if r.status_code == 200:
            data = r.json()
            if data.get("data"):
                return data["data"][0]["id"]
        else:
            logging.warning("GET /v1/models failed: %s %s", r.status_code, r.text[:200])
    except Exception as exc:
        logging.warning("GET /v1/models error: %s", exc)
    return None


###############################################################################
# Single benchmark point runner
###############################################################################

async def _run_point_async(
    *,
    api_url: str,
    model_id: str,
    api_key: str,
    input_requests: List[Tuple[str, int, int, None]],
    output_len: int,
    max_concurrency: int,
    goodput_config_dict: Dict[str, float],
    disable_pbar: bool = False,
) -> Tuple[BenchmarkMetrics, Dict[str, float], float]:
    semaphore = asyncio.Semaphore(max_concurrency)
    pbar = None
    if not disable_pbar and _TQDM_AVAILABLE:
        pbar = async_tqdm(total=len(input_requests), leave=False)

    t0 = time.perf_counter()
    tasks: List[asyncio.Task] = []
    async for req in get_request(input_requests, request_rate=float("inf")):
        prompt, prompt_len, out_len, _ = req
        inp = RequestFuncInput(
            model=model_id,
            prompt=prompt,
            api_url=api_url,
            prompt_len=prompt_len,
            output_len=out_len,
            ignore_eos=True,
            temperature=0.0,
        )
        tasks.append(
            asyncio.create_task(
                limited_request_func(
                    semaphore=semaphore,
                    request_func_input=inp,
                    api_key=api_key,
                    enable_thinking=False,
                    pbar=pbar,
                )
            )
        )
    outputs = await asyncio.gather(*tasks)
    if pbar is not None:
        pbar.close()
    dur_s = time.perf_counter() - t0

    metrics, _, per_key_pct = calculate_metrics(
        outputs=tuple(outputs),
        dur_s=dur_s,
        selected_percentiles=_SELECTED_PERCENTILES,
        goodput_config_dict=goodput_config_dict,
    )
    return metrics, per_key_pct, dur_s


def run_one_point(
    cfg: SweepConfig,
    model_id: str,
    tokenizer: Optional[Any],
    seqlen: int,
    batch_size: int,
    goodput_config_dict: Dict[str, float],
) -> Dict[str, Any]:
    """Run a single (seqlen, batch_size) benchmark and return a flat result dict.

    num_prompts = batch_size × combo_per_batch_size
    """
    num_prompts = batch_size * cfg.combo_per_batch_size
    api_url = cfg.base_url.rstrip("/") + "/v1/chat/completions"

    input_requests = make_input_requests(
        prompt_mode=cfg.prompt_mode,
        tokenizer=tokenizer,
        num_prompts=num_prompts,
        input_len=seqlen,
        output_len=cfg.output_len,
        seed=cfg.seed,
    )

    metrics, per_key_pct, dur_s = asyncio.run(
        _run_point_async(
            api_url=api_url,
            model_id=model_id,
            api_key=cfg.api_key,
            input_requests=input_requests,
            output_len=cfg.output_len,
            max_concurrency=batch_size,
            goodput_config_dict=goodput_config_dict,
        )
    )

    pct_ttft = metrics.percentiles_ttft_ms
    pct_tpot = metrics.percentiles_tpot_ms
    pct_itl = metrics.percentiles_itl_ms
    pct_e2el = metrics.percentiles_e2el_ms
    pct_dec = metrics.percentiles_decoding_throughput
    pct_pre = metrics.percentiles_prefilling_throughput

    row: Dict[str, Any] = {
        "seqlen": seqlen,
        "output_len": cfg.output_len,
        "batch_size": batch_size,
        "num_prompts": num_prompts,
        "duration_s": round(dur_s, 3),
        "completed": metrics.completed,
        "mean_input_tokens": metrics.mean_input,
        "mean_output_tokens": metrics.mean_output,
        # combined goodput
        "goodput_percentage": round(metrics.goodput_percentage, 4),
        # per-key goodput（e.g. ttft_goodput_pct, throughput_goodput_pct）
        **{
            f"{key}_goodput_pct": round(pct, 4)
            for key, pct in per_key_pct.items()
        },
        # TTFT
        "mean_TTFT": round(metrics.mean_ttft_ms, 3),
        "median_TTFT": round(metrics.median_ttft_ms, 3),
        "std_TTFT": round(metrics.std_ttft_ms, 3),
        "TTFT_P90": round(pct_ttft[0][1], 3) if len(pct_ttft) > 0 else None,
        "TTFT_P95": round(pct_ttft[1][1], 3) if len(pct_ttft) > 1 else None,
        "TTFT_P99": round(pct_ttft[2][1], 3) if len(pct_ttft) > 2 else None,
        # TPOT / ITL / E2EL
        "mean_TPOT": round(metrics.mean_tpot_ms, 3),
        "mean_ITL": round(metrics.mean_itl_ms, 3),
        "mean_E2EL": round(metrics.mean_e2el_ms, 3),
        "E2EL_P99": round(pct_e2el[2][1], 3) if len(pct_e2el) > 2 else None,
        # Throughput
        "request_throughput": round(metrics.request_throughput, 4),
        "output_throughput": round(metrics.output_throughput, 4),
        "mean_decode_throughput": round(metrics.mean_decoding_throughput, 4),
        "P90_decode_throughput": round(pct_dec[0][1], 4) if len(pct_dec) > 0 else None,
        "mean_prefill_throughput": round(metrics.mean_prefilling_throughput, 4),
        # Goodput mean metrics
        "mean_goodput_TTFT_ms": (
            round(metrics.mean_goodput_ttft * 1000, 3)
            if metrics.mean_goodput_ttft == metrics.mean_goodput_ttft else None
        ),
        "mean_goodput_throughput": (
            round(metrics.mean_goodput_throughput, 4)
            if metrics.mean_goodput_throughput == metrics.mean_goodput_throughput else None
        ),
    }
    return row


###############################################################################
# Warmup
###############################################################################

async def _warmup_async(
    *,
    api_url: str,
    model_id: str,
    api_key: str,
    input_len: int,
    output_len: int,
    prompt_mode: str,
    tokenizer: Optional[Any],
    seed: int,
) -> None:
    rng = random.Random(seed ^ 0x9E3779B9)
    warm_len = max(1, min(32, input_len))
    if prompt_mode == "exact":
        assert tokenizer is not None
        prompt, plen = build_exact_prompt_tokens(tokenizer, warm_len, rng)
    else:
        prompt = build_approx_prompt_tokens(warm_len, rng)
        plen = warm_len
    inp = RequestFuncInput(
        model=model_id,
        prompt=prompt,
        api_url=api_url,
        prompt_len=plen,
        output_len=max(8, min(64, output_len)),
        ignore_eos=True,
    )
    res = await async_request_openai(inp, api_key=api_key, pbar=None)
    if not res.success:
        raise RuntimeError(f"Warmup failed: {res.error}")


###############################################################################
# Step-up concurrency sequence generator
###############################################################################

def generate_concurrency_steps(
    min_bs: int,
    max_bs: int,
    linear_step: int = 4,
) -> List[int]:
    """Generate the step-up concurrency sequence used for searching.

    Strategy:
      - Starting from min_bs, double the value while we are below the
        "doubling threshold" (first power of 2 that is >= 2 × linear_step,
        e.g. 8 when linear_step=4).
      - After crossing the threshold, add linear_step each time.
      - Always include min_bs and max_bs.

    Example (min_bs=1, linear_step=4, max_bs=64):
      1, 2, 4, 8, 12, 16, 20, 24, 28, 32, 36, 40, 44, 48, 52, 56, 60, 64
    """
    # Doubling threshold: first power of 2 >= 2 * linear_step
    threshold = 1
    while threshold < 2 * linear_step:
        threshold *= 2
    # e.g. linear_step=4 → threshold=8

    result: Set[int] = set()
    # Doubling phase: powers of 2 from min_bs up to threshold
    v = min_bs
    while v <= max_bs:
        result.add(v)
        if v >= threshold:
            break
        v = v * 2

    # Linear phase from threshold (or wherever we stopped), step = linear_step
    start_linear = max(threshold, min_bs)
    # Round up to next multiple of linear_step
    if start_linear % linear_step != 0:
        start_linear = ((start_linear // linear_step) + 1) * linear_step
    else:
        start_linear = start_linear + linear_step

    v = start_linear
    while v <= max_bs:
        result.add(v)
        v += linear_step

    # Always include both ends
    result.add(min_bs)
    result.add(max_bs)

    return sorted(x for x in result if min_bs <= x <= max_bs)


###############################################################################
# Step-up search: find max concurrency for EACH goodput key
###############################################################################

def _is_100pct(value: Any) -> bool:
    try:
        return float(value) >= 100.0
    except (TypeError, ValueError):
        return False


def find_all_max_concurrencies(
    cfg: SweepConfig,
    model_id: str,
    tokenizer: Optional[Any],
    seqlen: int,
    goodput_config_dict: Dict[str, float],
    partial_path: Path,
) -> Tuple[Dict[str, int], List[Dict[str, Any]]]:
    """Step-up search to find max concurrency for EACH goodput key independently.

    Tracked keys:
      "combined"   → goodput_percentage >= 100%
      "ttft"       → ttft_goodput_pct >= 100%   (only if in goodput_config_dict)
      "tpot"       → tpot_goodput_pct >= 100%
      "e2el"       → e2el_goodput_pct >= 100%
      "throughput" → throughput_goodput_pct >= 100%

    Search stops as soon as ALL tracked keys have each failed once.
    The max concurrency for a key = the last batch_size where it passed 100%.
    A key that never passed even at min_batch_size gets max_concurrency = 0.

    Returns:
        max_concurrencies  – {key: max_batch_size}
        tested_points      – list of result dicts for every tested batch_size
                             (for detail Excel)
    """
    step_seq = generate_concurrency_steps(cfg.min_batch_size, cfg.max_batch_size, cfg.linear_step)

    # Keys to track
    tracked_keys: List[str] = ["combined"] + sorted(goodput_config_dict.keys())
    active_keys: Set[str] = set(tracked_keys)
    last_passing_bs: Dict[str, int] = {k: 0 for k in tracked_keys}
    max_concurrencies: Dict[str, int] = {}
    tested_points: List[Dict[str, Any]] = []

    logging.info(
        "  [search] step_seq=%s  tracked_keys=%s",
        step_seq[:12],
        tracked_keys,
    )

    for bs in step_seq:
        if not active_keys:
            logging.info(
                "  [stop] all keys have failed → stopping at bs=%d (not tested)", bs
            )
            break

        logging.info(
            "  ─────────────────────────────────────────────────────────"
        )
        logging.info(
            "  [bench] seqlen=%d  bs=%d  num_prompts=%d  active=%s",
            seqlen, bs, bs * cfg.combo_per_batch_size, sorted(active_keys),
        )

        try:
            res = run_one_point(
                cfg=cfg,
                model_id=model_id,
                tokenizer=tokenizer,
                seqlen=seqlen,
                batch_size=bs,
                goodput_config_dict=goodput_config_dict,
            )
        except Exception as exc:
            logging.error("  [error] seqlen=%d bs=%d : %s", seqlen, bs, exc)
            traceback.print_exc()
            res = {
                "seqlen": seqlen,
                "output_len": cfg.output_len,
                "batch_size": bs,
                "num_prompts": bs * cfg.combo_per_batch_size,
                "error": str(exc),
                "goodput_percentage": 0.0,
                **{f"{k}_goodput_pct": 0.0 for k in goodput_config_dict},
            }

        tested_points.append(res)
        _append_partial(partial_path, res)

        # Log per-key results
        gp_combined = res.get("goodput_percentage", 0.0)
        key_pcts = {
            k: res.get(f"{k}_goodput_pct", res.get("goodput_percentage", 0.0))
            for k in goodput_config_dict
        }
        logging.info(
            "    combined=%.1f%%  %s",
            gp_combined,
            "  ".join(f"{k}={v:.1f}%" for k, v in key_pcts.items()),
        )

        # Update per-key tracking
        newly_failed: Set[str] = set()
        for key in list(active_keys):
            if key == "combined":
                pct = res.get("goodput_percentage", 0.0)
            else:
                pct = res.get(f"{key}_goodput_pct", 0.0)

            if _is_100pct(pct):
                last_passing_bs[key] = bs
            else:
                max_concurrencies[key] = last_passing_bs[key]
                newly_failed.add(key)
                logging.info(
                    "    → key=%r  FAILED at bs=%d  (max_concurrency=%d)",
                    key, bs, last_passing_bs[key],
                )

        active_keys -= newly_failed

    # Keys still active after exhausting the sequence passed at max tested bs
    for key in active_keys:
        max_concurrencies[key] = last_passing_bs[key]
        logging.info(
            "  [end] key=%r still active → max_concurrency=%d",
            key, last_passing_bs[key],
        )

    return max_concurrencies, tested_points


###############################################################################
# Output helpers
###############################################################################

def _append_partial(path: Path, record: Dict[str, Any]) -> None:
    with open(path, "a", encoding="utf-8") as f:
        f.write(json.dumps(record, ensure_ascii=False) + "\n")


def safe_model_stem(model_id: str) -> str:
    name = Path(str(model_id)).name.strip() or "model"
    name = re.sub(r'[<>:"/\\|?*\n\r\t]', "_", name)
    return (name.strip(" ._") or "model")[:120]


def _auto_col_width(ws: Any) -> None:
    for col_cells in ws.columns:
        length = max(
            (len(str(c.value)) for c in col_cells if c.value is not None), default=8
        )
        ws.column_dimensions[
            get_column_letter(col_cells[0].column)
        ].width = min(length * 1.12 + 2, 52)


_HEADER_FILL = None
_COLOR_FILLS: List[Any] = []
_THIN_BORDER = None
_BOLD_FONT = None


def _init_excel_styles() -> None:
    global _HEADER_FILL, _COLOR_FILLS, _THIN_BORDER, _BOLD_FONT
    if _HEADER_FILL is not None:
        return
    _HEADER_FILL = PatternFill(
        start_color="FFFF00", end_color="FFFF00", fill_type="solid"
    )
    _COLOR_FILLS = [
        PatternFill(start_color="C5D9F1", end_color="C5D9F1", fill_type="solid"),
        PatternFill(start_color="F2DCDB", end_color="F2DCDB", fill_type="solid"),
        PatternFill(start_color="EBF1DE", end_color="EBF1DE", fill_type="solid"),
        PatternFill(start_color="FDE9D9", end_color="FDE9D9", fill_type="solid"),
    ]
    _THIN_BORDER = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    _BOLD_FONT = Font(bold=True)


def _write_excel_from_rows(
    path: Path,
    rows: List[Dict[str, Any]],
    columns: List[str],
    color_by_col: Optional[str] = None,
) -> None:
    """Generic helper: write rows to an Excel sheet with header + alternating colors."""
    if not _EXCEL_AVAILABLE:
        logging.warning("openpyxl/pandas not available; skipping Excel: %s", path)
        return
    _init_excel_styles()

    # Collect any extra columns not in the fixed list
    extra = [c for c in (rows[0] if rows else {}) if c not in columns]
    all_cols = columns + extra
    curated = [{c: row.get(c, "") for c in all_cols} for row in rows]
    df = pd.DataFrame(curated, columns=all_cols)

    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active

    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)

    color_col_idx: Optional[int] = None
    color_map: Dict[Any, int] = {}
    counter = 0

    for row_idx, row in enumerate(ws.iter_rows(), start=1):
        if row_idx == 1:
            # Find color_by_col index
            for ci, cell in enumerate(row):
                cell.fill = _HEADER_FILL
                cell.border = _THIN_BORDER
                cell.font = _BOLD_FONT
                if color_by_col and cell.value == color_by_col:
                    color_col_idx = ci
        else:
            label = row[color_col_idx].value if color_col_idx is not None else None
            if label not in color_map:
                color_map[label] = counter % len(_COLOR_FILLS)
                counter += 1
            fill = _COLOR_FILLS[color_map[label]]
            for cell in row:
                cell.fill = fill
                cell.border = _THIN_BORDER
                cell.font = _BOLD_FONT

    _auto_col_width(ws)
    wb.save(path)


# Detail Excel columns for each seqlen (one row per tested batch_size)
_DETAIL_COLS: List[str] = [
    "batch_size", "num_prompts",
    "goodput_percentage",
    # per-key goodput pct injected dynamically
    "mean_TTFT", "median_TTFT", "TTFT_P90", "TTFT_P95", "TTFT_P99",
    "mean_TPOT", "mean_ITL", "mean_E2EL", "E2EL_P99",
    "mean_decode_throughput", "P90_decode_throughput",
    "output_throughput", "mean_prefill_throughput",
    "request_throughput",
    "mean_goodput_TTFT_ms", "mean_goodput_throughput",
    "completed", "duration_s", "seqlen", "output_len",
]

# Summary Excel columns (one row per seqlen)
_SUMMARY_COLS_BASE: List[str] = [
    "seqlen", "output_len",
    "max_bs_combined",
    # per-key max_bs injected dynamically: max_bs_ttft, max_bs_throughput, etc.
]


def write_detail_excel(
    path: Path,
    rows: List[Dict[str, Any]],
    goodput_keys: List[str],
) -> None:
    """Write per-seqlen detail Excel: one row per tested batch_size."""
    per_key_pct_cols = [f"{k}_goodput_pct" for k in sorted(goodput_keys)]
    # Insert per-key pct columns right after goodput_percentage
    idx = _DETAIL_COLS.index("goodput_percentage") + 1
    columns = _DETAIL_COLS[:idx] + per_key_pct_cols + _DETAIL_COLS[idx:]
    _write_excel_from_rows(path, rows, columns, color_by_col="batch_size")
    logging.info("[detail excel] %s  (%d rows)", path, len(rows))


def write_summary_excel(
    path: Path,
    rows: List[Dict[str, Any]],
    goodput_keys: List[str],
) -> None:
    """Write final summary Excel: one row per seqlen, columns = max_bs per key."""
    per_key_max_cols = [f"max_bs_{k}" for k in sorted(goodput_keys)]
    columns = _SUMMARY_COLS_BASE + per_key_max_cols
    _write_excel_from_rows(path, rows, columns, color_by_col="seqlen")
    logging.info("[summary excel] %s  (%d rows)", path, len(rows))


###############################################################################
# Main
###############################################################################

def main() -> int:
    parser = argparse.ArgumentParser(
        description=(
            "Step-up sweep: find max concurrency per goodput-key (100% threshold) "
            "for each seqlen × output_len workload."
        )
    )
    parser.add_argument(
        "--config",
        required=True,
        help="Path to config JSON (see goodput_sweep_config.example.json).",
    )
    args = parser.parse_args()

    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s %(levelname)s: %(message)s",
    )

    cfg = SweepConfig.from_file(args.config)
    goodput_config_dict = check_goodput_args(cfg.goodput)
    goodput_keys = sorted(goodput_config_dict.keys())  # e.g. ["throughput", "ttft"]

    output_path = Path(cfg.output_path).resolve()
    output_path.parent.mkdir(parents=True, exist_ok=True)
    partial_path = output_path.with_suffix(output_path.suffix + ".partial.jsonl")
    if partial_path.exists():
        partial_path.unlink()

    detail_dir = Path(cfg.detail_excel_dir).resolve() if cfg.detail_excel_dir else Path.cwd()
    detail_dir.mkdir(parents=True, exist_ok=True)

    # Model name
    api_key = cfg.api_key or os.environ.get("OPENAI_API_KEY", "EMPTY")
    model_id = cfg.model or ""
    if not model_id:
        model_id = fetch_model_name(cfg.base_url, api_key) or ""
    if not model_id:
        logging.error(
            "Cannot determine model name. Set 'model' in config or ensure /v1/models works."
        )
        return 1

    # Tokenizer (exact mode only)
    tokenizer: Optional[Any] = None
    if cfg.prompt_mode == "exact":
        if not _TRANSFORMERS_AVAILABLE:
            logging.error("transformers not installed; required for prompt_mode='exact'.")
            return 1
        tokenizer = AutoTokenizer.from_pretrained(cfg.tokenizer_path, trust_remote_code=True)

    seqlen_list = cfg.seqlen_list()
    step_seq_example = generate_concurrency_steps(
        cfg.min_batch_size, cfg.max_batch_size, cfg.linear_step
    )
    ts = time.strftime("%Y%m%d_%H%M%S", time.localtime())
    model_stem = safe_model_stem(model_id)

    logging.info("=" * 72)
    logging.info("[config] base_url           = %s", cfg.base_url)
    logging.info("[config] model              = %s", model_id)
    logging.info("[config] seqlen_list        = %s", seqlen_list)
    logging.info("[config] output_len         = %d", cfg.output_len)
    logging.info("[config] concurrency_steps  = %s", step_seq_example)
    logging.info("[config] combo_per_bs       = %d", cfg.combo_per_batch_size)
    logging.info("[config] goodput SLO        = %s", cfg.goodput)
    logging.info("[config] tracked keys       = combined + %s", goodput_keys)
    logging.info("[config] output_path        = %s", output_path)
    logging.info("[config] detail_excel_dir   = %s", detail_dir)
    logging.info("=" * 72)

    # Warmup
    if not cfg.disable_warmup:
        logging.info("[warmup] sending 1 warmup request …")
        api_url = cfg.base_url.rstrip("/") + "/v1/chat/completions"
        try:
            asyncio.run(
                _warmup_async(
                    api_url=api_url,
                    model_id=model_id,
                    api_key=api_key,
                    input_len=cfg.warmup_input_len,
                    output_len=cfg.warmup_output_len,
                    prompt_mode=cfg.prompt_mode,
                    tokenizer=tokenizer,
                    seed=cfg.seed,
                )
            )
            logging.info("[warmup] done.")
        except Exception as exc:
            logging.warning("[warmup] failed (continuing): %s", exc)

    summary_rows: List[Dict[str, Any]] = []
    total_seqlens = len(seqlen_list)

    for idx, seqlen in enumerate(seqlen_list, start=1):
        print(
            f"\n{'=' * 72}\n"
            f"[sweep] seqlen={seqlen}  output_len={cfg.output_len}  "
            f"({idx}/{total_seqlens})\n"
            f"{'=' * 72}",
            flush=True,
        )

        max_concurrencies, tested_points = find_all_max_concurrencies(
            cfg=cfg,
            model_id=model_id,
            tokenizer=tokenizer,
            seqlen=seqlen,
            goodput_config_dict=goodput_config_dict,
            partial_path=partial_path,
        )

        # ── per-seqlen detail Excel ────────────────────────────────────────
        if not cfg.disable_excel and tested_points:
            detail_path = (
                detail_dir
                / f"{model_stem}_seqlen{seqlen}_out{cfg.output_len}_{ts}.xlsx"
            )
            write_detail_excel(detail_path, tested_points, goodput_keys)

        # ── summary row ───────────────────────────────────────────────────
        summary_row: Dict[str, Any] = {
            "seqlen": seqlen,
            "output_len": cfg.output_len,
            "max_bs_combined": max_concurrencies.get("combined", 0),
            **{
                f"max_bs_{k}": max_concurrencies.get(k, 0)
                for k in goodput_keys
            },
        }
        summary_rows.append(summary_row)

        # Print per-seqlen result
        key_strs = "  ".join(
            f"max_bs_{k}={max_concurrencies.get(k, 0)}" for k in goodput_keys
        )
        logging.info(
            "[result] seqlen=%d  max_bs_combined=%d  %s",
            seqlen,
            max_concurrencies.get("combined", 0),
            key_strs,
        )

    # ── Write final summary JSONL ─────────────────────────────────────────
    with open(output_path, "w", encoding="utf-8") as f:
        for row in summary_rows:
            f.write(json.dumps(row, ensure_ascii=False) + "\n")
    logging.info("[done] summary JSONL → %s  (%d rows)", output_path, len(summary_rows))
    logging.info("[done] partial detail → %s", partial_path)

    # ── Console summary table ─────────────────────────────────────────────
    col_headers = ["seqlen", "output_len", "max_bs_combined"] + [
        f"max_bs_{k}" for k in goodput_keys
    ]
    col_width = max(len(h) for h in col_headers) + 2
    header_line = "  ".join(h.ljust(col_width) for h in col_headers)
    sep_line = "  ".join("-" * col_width for _ in col_headers)
    print(f"\n{'=' * len(header_line)}")
    print(header_line)
    print(sep_line)
    for row in summary_rows:
        vals = [str(row.get(h, "")) for h in col_headers]
        print("  ".join(v.ljust(col_width) for v in vals))
    print(f"{'=' * len(header_line)}\n")

    # ── Final summary Excel ───────────────────────────────────────────────
    if not cfg.disable_excel:
        if not _EXCEL_AVAILABLE:
            logging.warning(
                "openpyxl/pandas not installed. Install them to enable Excel output."
            )
        else:
            if cfg.excel_path:
                xlsx_path = Path(cfg.excel_path)
            else:
                xlsx_path = output_path.parent / f"{model_stem}_goodput_summary_{ts}.xlsx"
            write_summary_excel(xlsx_path, summary_rows, goodput_keys)
            logging.info("[done] summary Excel → %s", xlsx_path.resolve())

    return 0


if __name__ == "__main__":
    sys.exit(main())
